import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import time

# Настройки для отображения колонок в пандас
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.max_colwidth', None)


def find_coordinate_value_on_column(sheet_d,column: str, text) :
    """
    Функция для поиска координат ячейки с заданным текстом
    param sheet_d: объект листа где нужно искать
    param column : название колонки
    param text : значение ячейки координаты которой нужно найти.
    """

    result = None
    # перебираем значения в колонке пока не найдем нужный результат, иначе None
    for cell in sheet_d[column]:
        if cell.value == text:
            result = cell.coordinate

    # Преобразовываем координаты в кортеж вида (column,row)
    return openpyxl.utils.cell.coordinate_from_string(result)




def defining_boundaries_table(wb_d: openpyxl.Workbook):
    """
    Функция для определения границ извлекаемых данных
    Где левое верхнее поле это поле в первом столбце со значением 1
    А правое нижнее пересечение строки ФК.00 и колонки Вар.часть.
    """
    # Переключаемся на лист План
    sheet = wb_d['План']
    # Находим адрес левого верхнего края
    top_left_tuple = find_coordinate_value_on_column(sheet,'B','1')
    print(top_left_tuple)
    # находим адрес левого нижнего края
    bottom_left_tuple = find_coordinate_value_on_column(sheet,'B','ФК.00')
    print(bottom_left_tuple)
    #Получаем буквенное обозначение последней колонки
    letter_max_column = openpyxl.utils.get_column_letter(sheet.max_column)
    # Получаем номер конечной строки, там где появляется слово ФК.00
    number_max_row = bottom_left_tuple[1]
    # получаем координаты крайних ячеек для проведения среза
    top_left = f'{top_left_tuple[0]}{top_left_tuple[1]}'
    bottom_right = f'{letter_max_column}{bottom_left_tuple[1]}'
    # Получаем срез нужных нам данных
    print(bottom_right)
    # Получаем список кортежей, где каждая строка таблицы это кортеж со значениями
    data_disciplines = sheet[top_left:bottom_right]
    print(data_disciplines[1])
    # Пробуем создать датафрейм
    df = pd.DataFrame(columns=range(100))
    count_columns =0
    # перебираем список кортежей-строк
    for row in data_disciplines:
        # получаем список значений всех ячеек в строке
        lst_to_df = [cell.value for cell in row]
        # Создаем промежуточный датафрейм из листа
        template_df = pd.DataFrame(lst_to_df)
        # разворачиваем столбец в стркоу
        flat_df = template_df.transpose()
        # добавляем в базовый
        df = pd.concat([df,flat_df],ignore_index=True)


    print(df.shape)
    df.to_excel('t.xlsx')


# Подготавливаем данные
name_curriculum = 'data/УП ПМ 2021.xlsx'
path_to_end_folder = 'output'
t = time.localtime()
current_time = time.strftime('%H_%M_%S', t)

# Открываем документ
wb = openpyxl.load_workbook(name_curriculum, data_only=True)


data_disciplines = defining_boundaries_table(wb)




